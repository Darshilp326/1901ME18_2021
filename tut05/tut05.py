import os
import csv
import openpyxl

# Dictionary with roll no as key
names_dict = {}
# Dictionary with sub code as key
sub_dict = {}
grade_map = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7, 'CC': 6, 'CD': 5, 'DD': 4, 'F': 0, 'I': 0,'AA*': 10, 'AB*': 9, 'BB*': 8, 'BC*': 7, 'CC*': 6, 'CD*': 5, 'DD*': 4, 'F*': 0, 'I*': 0}

# Read names-roll.csv and make file of each roll-no
def readNames():
    file = open('names-roll.csv')
    reader = csv.reader(file)    

    outputPath = r'./output'
    if not os.path.exists(outputPath):
        os.makedirs(outputPath)

    for line in reader:
        roll = line[0]
        name = line[1]
        rollPath = 'output/'+roll+'.xlsx'
        if roll == 'Roll':
            continue
        else:
            names_dict[roll] = name
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Overall'
            sheet.append(['Roll No', roll])
            sheet.append(['Name', name])
            sheet.append(['Discipline',roll[4:6]])
        wb.save(rollPath)
    return 
   
# Read subjects-master.csv and insert in sub_dict for subject
def readSubjects():
    file = open('subjects_master.csv')
    reader = csv.reader(file)

    for line in reader:
        sub = line[0]
        name = line[1]
        ltp = line[2]
        creds = line[3]
        temp = [name, ltp, creds]
        if sub == 'subno':
            continue
        else:
            sub_dict[sub] = temp
    return

def generate_marksheet():    
    readNames()
    readSubjects()

    file = open('grades.csv')
    reader = csv.reader(file)
    lastRoll = ""
    spi = 0
    i = 0
    totalCredits = 0
    cpi = 0
    rollno=''
    sem_no_list = ['Semester']
    cpi_list = ['CPI']
    spi_list = ['SPI']
    sem_cred_list = ['Semester wise Credit Taken']
    total_credits=['Total credits']

    for grade in reader:
        f = 0
        rollno = grade[0]
        if rollno == 'Roll':
            continue
        resultPath = 'output/'+rollno+'.xlsx'
        wb = openpyxl.load_workbook(resultPath)
        for sheet in wb.worksheets:
            if sheet.title == grade[1]:
                f = 1
                break

        # If roll no is changed,add all stored data to last roll no's overall sheet.        
        if lastRoll != '':
            if lastRoll != rollno:
                path = 'output/'+lastRoll+'.xlsx'
                wb1 = openpyxl.load_workbook(path)
                spi_list.append(round(spi/totalCredits,2))
                sem_cred_list.append(totalCredits)
                sem_no_list.append(semester)
                total_credits.append(int(total_credits[-1])+totalCredits)
                cpi_list.append(round(cpi/total_credits[-1],2))
                overall_sheet = wb1['Overall']
                overall_sheet.append(sem_no_list)
                overall_sheet.append(sem_cred_list)
                overall_sheet.append(spi_list)
                overall_sheet.append(total_credits)
                overall_sheet.append(cpi_list)
                spi_list=['SPI']
                sem_cred_list=['Credits']
                sem_no_list=['Semester']
                total_credits=['Total Credits']
                cpi_list=['CPI']
                wb1.save(path)
                cpi=0
        lastRoll = rollno
        semester = grade[1]
        subcode = grade[2]
        credit = int(grade[3])
        marks = grade_map[grade[4].strip()]
        sub_type = grade[5]
        if f == 0:
            res = len(wb.sheetnames)
            sheet_wb2 = wb.create_sheet(semester)
            sheet_wb2.title = semester
            i = 0
            sheet_wb2.append(
                ['SI No', 'Code', 'Subject Name', 'LTP', 'Credits', 'Type', 'Grade'])

            if res > 1:
                spi_list.append(round(spi/totalCredits,2))
                sem_cred_list.append(totalCredits)
                sem_no_list.append(str(int(semester)-1))
                if(len(total_credits)==1):
                    total_credits.append(totalCredits)
                else:
                    total_credits.append(int(total_credits[-1])+totalCredits)
                cpi_list.append(round(cpi/total_credits[-1],2))        
            spi = 0
            totalCredits = 0
        else:
            sheet_wb2 = wb[semester]
        i=i+1
        spi+=marks*credit
        totalCredits+=credit
        cpi+=marks*credit
        sheet_wb2.append(
            [i, subcode, sub_dict[subcode][0], sub_dict[subcode][1], sub_dict[subcode][2], sub_type, grade[4]])    
        wb.save(resultPath)

    # Last roll's overall sheet has to be filled 
    path = 'output/'+rollno+'.xlsx'
    wb1 = openpyxl.load_workbook(path)
    spi_list.append(round(spi/totalCredits,2))
    sem_cred_list.append(totalCredits)
    sem_no_list.append(semester)
    total_credits.append(int(total_credits[-1])+totalCredits)
    cpi_list.append(round(cpi/total_credits[-1],2))
    overall_sheet = wb1['Overall']
    overall_sheet.append(sem_no_list)
    overall_sheet.append(sem_cred_list)
    overall_sheet.append(spi_list)
    overall_sheet.append(total_credits)
    overall_sheet.append(cpi_list)
    wb1.save(path)
    return

generate_marksheet()