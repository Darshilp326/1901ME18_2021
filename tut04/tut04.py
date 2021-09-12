import os
import csv
import openpyxl


def output_by_subject():
    file = open('regtable_old.csv')
    csvreader = csv.reader(file)
    f = []
    for line in csvreader:
        f.append(line)
    data = []
    subPath = r'./output_by_subject' 
    if not os.path.exists(subPath):
        os.makedirs(subPath)

    for line in f:
        rollno = line[0]
        registerSem = line[1]
        subNo = line[3]
        subType = line[8]

        path = 'output_by_subject/' + subNo + '.xlsx'

        if(rollno == 'rollno'):
            continue
        if(subNo in data):
            path = 'output_by_subject/' + subNo + '.xlsx'
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            sheet.append([rollno, registerSem, subNo, subType])

        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            data.append(subNo)
            sheet.append(["rollno","register_sem","sub_no","sub_type"])
            sheet.append([rollno, registerSem, subNo, subType])

        wb.save(path)

    return

def output_individual_roll():
    file = open('regtable_old.csv')
    csvreader = csv.reader(file)
    lines = []
    for line in csvreader:
        lines.append(line)
    lines.sort()
    result = {}
    last_roll = ""
    for line in lines:
        words = line
        if words[0] == 'rollno':
            continue
        if last_roll != words[0]:
            subjectList = []
        subjectList.append(
            {"sem": words[1], "sub_no": words[3], "sub_type": words[8]})
        result[words[0]] = subjectList
        last_roll = words[0]

    rollPath = r'./output_individual_roll' 
    if not os.path.exists(rollPath):
        os.makedirs(rollPath)

    for x in result:
        filepath = "output_individual_roll/%s.xlsx" % x
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["rollno","register_sem","sub_no","sub_type"])
        for b in result[x]:
            sheet.append([x, b['sem'], b['sub_no'], b['sub_type']])
        wb.save(filepath)
    return


output_individual_roll()
output_by_subject()